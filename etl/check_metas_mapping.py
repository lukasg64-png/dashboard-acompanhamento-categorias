import openpyxl, json, os

dst = 'temp_novas_metas.xlsx'
wb = openpyxl.load_workbook(dst, data_only=True)
ws = wb.active

excel_distritos = set()
excel_linhas = set()
total_meta_excel = 0

rows = []
for r in range(4, ws.max_row + 1):
    dist = ws.cell(r, 1).value
    linha = ws.cell(r, 2).value
    familia = ws.cell(r, 3).value
    meta = ws.cell(r, 4).value
    if dist and linha and meta is not None:
        try:
            val = float(meta)
            total_meta_excel += val
            excel_distritos.add(str(dist).strip())
            excel_linhas.add(str(linha).strip())
            rows.append({
                'distrito': str(dist).strip(),
                'linha': str(linha).strip(),
                'familia': str(familia).strip() if familia else '',
                'meta': val
            })
        except:
            pass

print(f"Total rows: {len(rows)}")
print(f"Total Meta Excel: R$ {total_meta_excel:,.2f}")
print(f"Unique Distritos in Excel ({len(excel_distritos)}): {sorted(list(excel_distritos))}")
print(f"Unique Linhas in Excel ({len(excel_linhas)})")

# Check mapping with filtro_hierarquia.json
hier_path = 'data/setembro/filtro_hierarquia.json'
with open(hier_path, 'r', encoding='utf-8') as f:
    hier = json.load(f)

print(f"\nKeys in filtro_hierarquia.json: {list(hier.keys())}")
for k, v in hier.items():
    if isinstance(v, list):
        print(f"  {k} (list len={len(v)}): {v[:5]}")
    elif isinstance(v, dict):
        print(f"  {k} (dict len={len(v)}): keys={list(v.keys())[:5]}")

# Check sample from hierarquia_detalhada.json
det_path = 'data/setembro/hierarquia_detalhada.json'
with open(det_path, 'r', encoding='utf-8') as f:
    det = json.load(f)

print(f"\nHierarquia detalhada items count: {len(det)}")
if len(det) > 0:
    print(f"Sample row keys: {list(det[0].keys())}")
    print(f"Sample row: {det[0]}")

# Check how distritais and diretores appear in hierarquia_detalhada
dir_dist_pairs = set()
linha_to_grupo_subgrupo = {}
for item in det:
    d = item.get('diretor')
    dist = item.get('distrital')
    if d or dist:
        dir_dist_pairs.add((d, dist))
    
    linha = item.get('linha')
    grupo = item.get('grupo')
    subgrupo = item.get('subgrupo')
    if linha and linha not in linha_to_grupo_subgrupo:
        linha_to_grupo_subgrupo[linha] = {'grupo': grupo, 'subgrupo': subgrupo}

print(f"\nDiretor -> Distrital pairs found in Qlik dataset ({len(dir_dist_pairs)}):")
for d, dist in sorted(list(dir_dist_pairs)):
    print(f"  Diretor: {d} -> Distrital: {dist}")

# Match distritos from Excel
excel_to_dir = {}
for dist in excel_distritos:
    matching_dirs = [d for (d, qdist) in dir_dist_pairs if qdist and (dist.lower() in qdist.lower() or qdist.lower() in dist.lower())]
    excel_to_dir[dist] = matching_dirs

print(f"\nExcel Distritos mapping to Diretores:")
for dist, dirs in excel_to_dir.items():
    print(f"  Distrito Excel '{dist}' -> Diretores Qlik: {dirs}")

# Match Linhas from Excel to Grupo/Subgrupo
matched_linhas = 0
unmatched_linhas = []
for row in rows:
    l = row['linha']
    if l in linha_to_grupo_subgrupo:
        matched_linhas += 1
    else:
        unmatched_linhas.append(l)

print(f"\nLinhas matched: {matched_linhas} / {len(rows)} ({matched_linhas/len(rows)*100:.1f}%)")
if unmatched_linhas:
    print(f"Unmatched linhas sample (first 10 of {len(set(unmatched_linhas))} unique): {list(set(unmatched_linhas))[:10]}")
