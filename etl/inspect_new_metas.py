import shutil, os, json
import openpyxl

src = 'novas metas por distrital.xlsx'
dst = 'temp_novas_metas.xlsx'

if not os.path.exists(dst) or os.path.getmtime(src) > os.path.getmtime(dst):
    shutil.copy2(src, dst)

wb = openpyxl.load_workbook(dst, data_only=True)
print("Sheet names:", wb.sheetnames)

for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n==========================================")
    print(f"Sheet: {name} (max_row={ws.max_row}, max_col={ws.max_column})")
    print(f"==========================================")
    for r in range(1, min(15, ws.max_row + 1)):
        row_vals = [ws.cell(r, c).value for c in range(1, min(25, ws.max_column + 1))]
        if any(v is not None for v in row_vals):
            print(f"Row {r:2d}: {row_vals}")
