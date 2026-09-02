import openpyxl, json, os, shutil
import pandas as pd
import numpy as np

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, 'novas metas por distrital.xlsx')
TEMP_EXCEL = os.path.join(BASE_DIR, 'temp_novas_metas.xlsx')
HIER_JSON = os.path.join(BASE_DIR, 'data', 'setembro', 'hierarquia_detalhada.json')
OUT_JSON = os.path.join(BASE_DIR, 'data', 'setembro', 'dashboard_setembro.json')

# Copy excel to avoid lock
if not os.path.exists(TEMP_EXCEL) or os.path.getmtime(EXCEL_PATH) > os.path.getmtime(TEMP_EXCEL):
    shutil.copy2(EXCEL_PATH, TEMP_EXCEL)

wb = openpyxl.load_workbook(TEMP_EXCEL, data_only=True)
ws = wb.active

# 1. Read Excel Distrital x Linha
# Headers: Distrito | Linha | Família | Meta Set/26
excel_rows = []
for r in range(4, ws.max_row + 1):
    dist = ws.cell(r, 1).value
    linha = ws.cell(r, 2).value
    familia = ws.cell(r, 3).value
    meta = ws.cell(r, 4).value
    if dist and linha and meta is not None:
        try:
            val = float(meta)
            excel_rows.append({
                'distrital': str(dist).strip(),
                'linha': str(linha).strip(),
                'familia': str(familia).strip() if familia else '',
                'meta_mes': val
            })
        except:
            pass

df_metas = pd.DataFrame(excel_rows)
print(f"Total rows in df_metas: {len(df_metas)}")
print(f"Total Meta Empresa: R$ {df_metas['meta_mes'].sum():,.2f}")

# 2. Read Qlik Realizado Setembro (D-1)
with open(HIER_JSON, 'r', encoding='utf-8') as f:
    hier_list = json.load(f)

df_real = pd.DataFrame(hier_list)
print(f"Total rows in df_real: {len(df_real)}")
print(f"Total Realizado Setembro D-1: R$ {df_real['venda_jul_26'].sum():,.2f}")

# Map Diretor to Distrital
dir_map = {}
for _, row in df_real.iterrows():
    d = row.get('diretor')
    dist = row.get('distrital')
    if d and dist and dist not in dir_map:
        dir_map[dist] = d

print("Diretor mapping:", dir_map)
df_metas['diretor'] = df_metas['distrital'].map(dir_map).fillna('Outros')

# Linha to Grupo & Subgrupo mapping from Qlik
linha_map = {}
for _, row in df_real.iterrows():
    l = row.get('linha')
    g = row.get('grupo')
    sg = row.get('subgrupo')
    if l and l not in linha_map:
        linha_map[l] = {'grupo': g, 'subgrupo': sg}

df_metas['grupo'] = df_metas['linha'].apply(lambda x: linha_map.get(x, {}).get('grupo', 'Outros'))
df_metas['subgrupo'] = df_metas['linha'].apply(lambda x: linha_map.get(x, {}).get('subgrupo', 'Outros'))

# Group Realizado by (diretor, distrital, grupo, subgrupo, linha)
df_real_agg = df_real.groupby(['diretor', 'distrital', 'grupo', 'subgrupo', 'linha']).agg(
    venda_set26=('venda_jul_26', 'sum'),
    venda_ago26=('venda_jun_26', 'sum'),
    venda_set25=('venda_jul_25', 'sum')
).reset_index()

# Merge Metas and Realizado
df_merged = pd.merge(
    df_metas,
    df_real_agg,
    on=['diretor', 'distrital', 'grupo', 'subgrupo', 'linha'],
    how='outer'
).fillna({
    'meta_mes': 0.0,
    'venda_set26': 0.0,
    'venda_ago26': 0.0,
    'venda_set25': 0.0,
    'diretor': 'Outros',
    'distrital': 'Outros',
    'grupo': 'Outros',
    'subgrupo': 'Outros',
    'linha': 'Outros'
})

print(f"Merged total rows: {len(df_merged)}")
print(f"Merged Meta Total: R$ {df_merged['meta_mes'].sum():,.2f}")
print(f"Merged Realizado Total: R$ {df_merged['venda_set26'].sum():,.2f}")

# Summary by Diretoria
print("\n--- RESUMO POR DIRETORIA ---")
df_dir = df_merged.groupby('diretor').agg(
    meta=('meta_mes', 'sum'),
    realizado=('venda_set26', 'sum')
).reset_index()
df_dir['atingimento'] = df_dir['realizado'] / (df_dir['meta'] / 30) * 100 # atingimento D1
for _, r in df_dir.iterrows():
    print(f"  Diretor: {r['diretor']:18s} | Meta Mês: R$ {r['meta']:12,.2f} | Realizado D1: R$ {r['realizado']:10,.2f} | Ating. D1: {r['atingimento']:5.1f}%")

# Summary by Distrital
print("\n--- RESUMO POR DISTRITAL ---")
df_dist = df_merged.groupby(['diretor', 'distrital']).agg(
    meta=('meta_mes', 'sum'),
    realizado=('venda_set26', 'sum')
).reset_index()
df_dist['atingimento'] = df_dist['realizado'] / (df_dist['meta'] / 30) * 100
for _, r in df_dist.iterrows():
    print(f"  Distrital: {r['distrital']:20s} ({r['diretor']:18s}) | Meta: R$ {r['meta']:11,.2f} | Real: R$ {r['realizado']:10,.2f} | Ating: {r['atingimento']:5.1f}%")

# Summary by Grupo
print("\n--- RESUMO POR GRUPO EMPRESA ---")
df_grp = df_merged.groupby('grupo').agg(
    meta=('meta_mes', 'sum'),
    realizado=('venda_set26', 'sum')
).reset_index()
df_grp['atingimento'] = df_grp['realizado'] / (df_grp['meta'] / 30) * 100
for _, r in df_grp.sort_values(by='meta', ascending=False).iterrows():
    print(f"  Grupo: {r['grupo']:25s} | Meta: R$ {r['meta']:12,.2f} | Real: R$ {r['realizado']:10,.2f} | Ating: {r['atingimento']:5.1f}%")
