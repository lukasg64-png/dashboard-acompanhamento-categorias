"""
load_metas_setembro.py — Processa o novo arquivo de metas de Setembro/2026:
"novas metas por distrital.xlsx" (Distrital × Linha × Família).
Gera dados de metas desdobrados por Diretoria, Distrital, Grupo e Linha.
"""
import os, sys, json, shutil
if hasattr(sys.stdout, 'reconfigure'): sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'): sys.stderr.reconfigure(encoding='utf-8')

import openpyxl
import pandas as pd

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, 'data', 'setembro')
EXCEL_PRIMARY = os.path.join(BASE_DIR, 'novas metas por distrital.xlsx')
EXCEL_TEMP = os.path.join(BASE_DIR, 'temp_novas_metas.xlsx')

# Mapeamento Oficial Diretor Regional ➔ Distrital
DIRETOR_MAP = {
    'Fais Abdalla': 'Cintia Silva',
    'Juliano Petrarca': 'Cintia Silva',
    'Rodrigo Ferreira': 'Cintia Silva',
    'Tiago Machado': 'Cintia Silva',
    'Andre Jaime': 'Laerti Siqueira',
    'Cleiton Lima': 'Laerti Siqueira',
    'Fabio Baldasso': 'Laerti Siqueira',
    'Larissa Azambuja': 'Laerti Siqueira',
    'Luiz Kuhn': 'Laerti Siqueira'
}

def get_excel_workbook():
    """Lê o arquivo Excel criando cópia temporária se necessário para evitar bloqueio."""
    try:
        shutil.copy2(EXCEL_PRIMARY, EXCEL_TEMP)
        target = EXCEL_TEMP
    except Exception:
        target = EXCEL_PRIMARY if os.path.exists(EXCEL_PRIMARY) else EXCEL_TEMP
    
    return openpyxl.load_workbook(target, data_only=True)

def load_curva_diaria_setembro():
    """Gera ou carrega a curva diária de Setembro/2026 (30 dias)."""
    curva_file = os.path.join(DATA_DIR, 'curva_diaria.json')
    if os.path.exists(curva_file):
        with open(curva_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    
    # Fallback linear se não existir
    dias = []
    pct_dia = 1.0 / 30.0
    for d in range(1, 31):
        dias.append({
            'dia': d,
            'dia_semana': '',
            'peso': 1.0,
            'pct_dia': pct_dia,
            'pct_acum': pct_dia * d,
            'proj_dia': 0,
            'proj_acum': 0,
            'meta_dia': 0,
            'meta_acum': 0
        })
    return dias

def main():
    os.makedirs(DATA_DIR, exist_ok=True)
    print("\n" + "=" * 70)
    print("  PROCESSANDO NOVAS METAS SETEMBRO/2026 (POR DISTRITAL & LINHA)")
    print("=" * 70)

    # 1. Carregar mapeamento de Linha -> Grupo/Subgrupo a partir de hierarquia_detalhada.json
    hier_json_path = os.path.join(DATA_DIR, 'hierarquia_detalhada.json')
    linha_to_grupo = {}
    linha_to_subgrupo = {}
    if os.path.exists(hier_json_path):
        with open(hier_json_path, 'r', encoding='utf-8') as f:
            hier_list = json.load(f)
            for item in hier_list:
                l = item.get('linha')
                g = item.get('grupo')
                sg = item.get('subgrupo')
                if l:
                    if g and l not in linha_to_grupo: linha_to_grupo[l] = g
                    if sg and l not in linha_to_subgrupo: linha_to_subgrupo[l] = sg

    # 2. Carregar Excel "novas metas por distrital.xlsx"
    wb = get_excel_workbook()
    ws = wb.active
    print(f"  Planilha ativa: {ws.title} ({ws.max_row} linhas)")

    rows = []
    for r in range(4, ws.max_row + 1):
        dist = ws.cell(r, 1).value
        linha = ws.cell(r, 2).value
        familia = ws.cell(r, 3).value
        meta = ws.cell(r, 4).value

        if dist and linha and meta is not None:
            try:
                meta_val = float(meta)
                dist_str = str(dist).strip()
                linha_str = str(linha).strip()
                familia_str = str(familia).strip() if familia else ''
                diretor_str = DIRETOR_MAP.get(dist_str, 'Outros')
                grupo_str = linha_to_grupo.get(linha_str, 'Outros')
                subgrupo_str = linha_to_subgrupo.get(linha_str, 'Outros')

                rows.append({
                    'diretor': diretor_str,
                    'distrital': dist_str,
                    'grupo': grupo_str,
                    'subgrupo': subgrupo_str,
                    'linha': linha_str,
                    'familia': familia_str,
                    'meta_mensal': meta_val
                })
            except Exception:
                pass

    df = pd.DataFrame(rows)
    raw_total = df['meta_mensal'].sum()
    print(f"  Total registros processados: {len(df)}")
    print(f"  Meta Raw Total Setembro: R$ {raw_total:,.4f}")

    # Arredondar para 2 casas e fazer balanceamento de centavos para bater 100.0000% com o Excel
    df['meta_mensal'] = df['meta_mensal'].round(2)
    diff_cents = int(round((raw_total - df['meta_mensal'].sum()) * 100))
    if diff_cents != 0:
        top_indices = df.nlargest(abs(diff_cents), 'meta_mensal').index
        step = 0.01 if diff_cents > 0 else -0.01
        for idx in top_indices:
            df.loc[idx, 'meta_mensal'] = round(df.loc[idx, 'meta_mensal'] + step, 2)

    meta_empresa_total = round(df['meta_mensal'].sum(), 2)
    print(f"  Meta Final Balanceada (Centavos Exatos): R$ {meta_empresa_total:,.2f}")

    # 3. Carregar curva diária para calcular metas diárias
    curva = load_curva_diaria_setembro()
    pcts_dia = [c['pct_dia'] for c in curva]

    # Calcular meta acumulada da empresa
    meta_empresa_total = df['meta_mensal'].sum()
    for c in curva:
        c['meta_dia'] = round(meta_empresa_total * c['pct_dia'], 2)
        c['meta_acum'] = round(meta_empresa_total * c['pct_acum'], 2)

    with open(os.path.join(DATA_DIR, 'curva_diaria.json'), 'w', encoding='utf-8') as f:
        json.dump(curva, f, ensure_ascii=False, indent=2)

    # 4. Consolidar por Linha Empresa (Macro)
    df_linhas = df.groupby(['linha', 'familia', 'grupo', 'subgrupo']).agg(
        meta_mensal=('meta_mensal', 'sum')
    ).reset_index()

    metas_por_linha_dia = []
    for _, r in df_linhas.iterrows():
        mm = r['meta_mensal']
        meta_diaria = [round(mm * p, 2) for p in pcts_dia]
        meta_acum = []
        ac = 0.0
        for md in meta_diaria:
            ac += md
            meta_acum.append(round(ac, 2))

        metas_por_linha_dia.append({
            'linha': r['linha'],
            'familia': r['familia'],
            'categoria': r['grupo'],
            'grupo': r['grupo'],
            'subgrupo': r['subgrupo'],
            'meta_mensal': mm,
            'meta_diaria': meta_diaria,
            'meta_acum': meta_acum
        })

    with open(os.path.join(DATA_DIR, 'metas_por_linha_dia.json'), 'w', encoding='utf-8') as f:
        json.dump(metas_por_linha_dia, f, ensure_ascii=False)
    print(f"  [OK] metas_por_linha_dia.json gerado ({len(metas_por_linha_dia)} linhas)")

    # 5. Consolidar por Distrital × Linha (Micro)
    metas_distrital_linha = []
    for _, r in df.iterrows():
        mm = r['meta_mensal']
        meta_diaria = [round(mm * p, 2) for p in pcts_dia]
        meta_acum = []
        ac = 0.0
        for md in meta_diaria:
            ac += md
            meta_acum.append(round(ac, 2))

        metas_distrital_linha.append({
            'diretor': r['diretor'],
            'distrital': r['distrital'],
            'grupo': r['grupo'],
            'subgrupo': r['subgrupo'],
            'linha': r['linha'],
            'familia': r['familia'],
            'meta_mensal': mm,
            'meta_diaria': meta_diaria,
            'meta_acum': meta_acum
        })

    with open(os.path.join(DATA_DIR, 'metas_distrital_linha.json'), 'w', encoding='utf-8') as f:
        json.dump(metas_distrital_linha, f, ensure_ascii=False)
    print(f"  [OK] metas_distrital_linha.json gerado ({len(metas_distrital_linha)} registros)")

    # 6. Gerar Resumo de Metas
    resumo = {
        'meta_empresa_total': round(meta_empresa_total, 2),
        'total_linhas': len(df_linhas),
        'total_distritais': df['distrital'].nunique(),
        'total_diretorias': df['diretor'].nunique(),
        'diretorias': {}
    }

    for dir_nome, grp_dir in df.groupby('diretor'):
        resumo['diretorias'][dir_nome] = {
            'meta_total': round(grp_dir['meta_mensal'].sum(), 2),
            'distritais': {}
        }
        for dist_nome, grp_dist in grp_dir.groupby('distrital'):
            resumo['diretorias'][dir_nome]['distritais'][dist_nome] = {
                'meta_total': round(grp_dist['meta_mensal'].sum(), 2),
                'total_linhas': len(grp_dist)
            }

    with open(os.path.join(DATA_DIR, 'resumo_metas.json'), 'w', encoding='utf-8') as f:
        json.dump(resumo, f, ensure_ascii=False, indent=2)
    print(f"  [OK] resumo_metas.json gerado")
    print("=" * 70)

if __name__ == '__main__':
    main()
